import tkinter
import openpyxl
import datetime
import tkinter.ttk
import pandas as pd
import tkinter.simpledialog
from unidecode import unidecode

class App(tkinter.Frame):
    def __init__(self, master):
        super().__init__(master)
        self.master.title("Instrumentos de Medição")
        self.master.minsize(width=300, height=200)
        self.master.maxsize(width=1300, height=500)

        self.pack()
        self.paginaInicial()

    def limpar(self):
        for widget in self.master.winfo_children():
            widget.destroy()
        self.master.minsize(width=500, height=200)
        self.master.maxsize(width=500, height=200)

    def paginaInicial(self):
        self.limpar()
        
        title = tkinter.Label(self.master, text='Instrumentos de Medição', font=("bold", 16))
        title.pack(fill=tkinter.X, pady=10)

        self.botao_norma = tkinter.Button(text='Norma',command=self.normaInput)
        self.botao_norma.place(relx=0.5, rely=0.8, anchor='s')

        self.botao_kw = tkinter.Button(text='KW',command=self.kwInput)
        self.botao_kw.place(relx=0.5, rely=0.65, anchor='s')

        self.botao_instru = tkinter.Button(text='Instrumento',command=self.instruInput)
        self.botao_instru.place(relx=0.5, rely=0.5, anchor='s')
        
    def normaInput(self):
        self.limpar()
        
        self.label_nome_instrumento = tkinter.Label(text='Norma: ')
        self.label_nome_instrumento.place(relx=0.3, rely=0.3, anchor='nw')

        self.entrada_nome_instrumento = tkinter.Entry(borderwidth=5, width=30)
        self.entrada_nome_instrumento.place(relx=0.45, rely=0.3, anchor='nw')

        self.nome_instrumento = tkinter.StringVar()
        self.entrada_nome_instrumento['textvariable'] = self.nome_instrumento

        self.botao_confirmacao = tkinter.Button(text='Confirmar',command=self.buscaNorma)
        self.botao_confirmacao.place(relx=0.5, rely=0.65, anchor='s')
        
        self.botao_voltar = tkinter.Button(text="Voltar", command = self.paginaInicial)
        self.botao_voltar.place(relx=0.005, rely=0.005, anchor='nw')

    def kwInput(self):
        self.limpar()

        self.label_kw = tkinter.Label(text='KW: ')
        self.label_kw.place(relx=0.3, rely=0.3, anchor='nw')

        self.entrada_kw = tkinter.Entry(borderwidth=5, width=30)
        self.entrada_kw.place(relx=0.45, rely=0.3, anchor='nw')
        
        self.kw = tkinter.StringVar()
        self.entrada_kw['textvariable'] = self.kw

        def valida():
            try:
                kw = self.kw.get()
                kw = kw.split('/')
                ano = int(kw[1])
                semana = int(kw[0])

                if semana > 52 or semana < 1  or ano >= 100:
                    raise ValueError() 
                else: 
                    self.buscaKW('kw')
            except (ValueError, IndexError):
                tkinter.messagebox.showerror(title='Erro',message='Valor inválido, tentar novamente.')

        self.botao_confirmacao = tkinter.Button(text='Confirmar',command=valida)
        self.botao_confirmacao.place(relx=0.5, rely=0.65, anchor='s')

        self.botao_voltar = tkinter.Button(text="Voltar", command = self.paginaInicial)
        self.botao_voltar.place(relx=0.005, rely=0.005, anchor='nw')

    def instruInput(self):
        self.limpar()

        self.label_instru = tkinter.Label(text='Instrumento: ')
        self.label_instru.place(relx=0.3, rely=0.3, anchor='nw')

        self.entrada_instru = tkinter.Entry(borderwidth=5, width=30)
        self.entrada_instru.place(relx=0.45, rely=0.3, anchor='nw')
        
        self.instru = tkinter.StringVar()
        self.entrada_instru['textvariable'] = self.instru

        self.botao_confirmacao = tkinter.Button(text='Confirmar',command=self.buscaInstrumento)
        self.botao_confirmacao.place(relx=0.5, rely=0.65, anchor='s')
        
        self.botao_voltar = tkinter.Button(text="Voltar", command = self.paginaInicial)
        self.botao_voltar.place(relx=0.005, rely=0.005, anchor='nw')

    def buscaNorma(self):
        def estadoAtual(planilha, col_cell, row_cell, workbook):
            if "Perdidos" in planilha:
                planilha.clear()
                planilha.append("Perdidos")
            elif "Sucateados" in planilha:
                planilha.clear()
                planilha.append("Sucateados")

        self.index = []
        nome_instrumento = self.nome_instrumento.get()
        nome_instrumento = nome_instrumento.strip()

        final = nome_instrumento.split('-')
        final = final[-1].strip()

        nome_instrumento = ''.join(nome_instrumento.split())

        try:
            workbook = openpyxl.load_workbook(filename='Instrumentos de medição.xlsx')
            planilha = []
            values = []
            for worksheets in workbook.sheetnames:
                worksheet = workbook[worksheets]
                for col in worksheet.iter_cols(min_col = 1, max_col = 8, min_row= 1, max_row= 1):
                    for col_cell in col:
                        if(col_cell.value == "NORMA/SÉRIE"):
                            for row in worksheet.iter_rows(min_col = col_cell.column, max_col = col_cell.column,
                                                           min_row = 2, max_row = worksheet.max_row):
                                for row_cell in row:
                                    if type(row_cell.value) == str:
                                        nome = row_cell.value.strip()
                                        compFinal = [nome]

                                        if '-' in nome:
                                            compFinal = nome.split('-')
                                            compFinal = compFinal[1].strip()

                                        if ''.join(nome.split()) == nome_instrumento or final == compFinal:
                                            planilha.append(worksheet.title)
                                            estadoAtual(planilha, col_cell, row_cell, workbook)
                                            worksheet = workbook[planilha[0]]
                                            self.index.append(row_cell.row)
                                            if row_cell.parent.title != planilha[0]:
                                                break
                                            for print_row in worksheet.iter_cols(min_col=worksheet.min_column,
                                                                                max_col=worksheet.max_column,
                                                                                min_row=row_cell.row,
                                                                                max_row=row_cell.row):
                                                for row in print_row:
                                                    values.append(row.value)
                                        break
                            break
            
            if len(values) > 8:
                values = values[8:]

            if planilha[0] == "Sucateados":
                colunas = ["LOCALIZAÇÃO", "LETRA",	"DENOMINAÇÃO",	"MARCA", "NORMA/SÉRIE"]	
                self.df = pd.DataFrame([values], columns=colunas)
            else:
                colunas = ["LOCALIZAÇÃO", "LETRA",	"DENOMINAÇÃO",	"MARCA", "NORMA/SÉRIE",	"DATA DE RETIRADA",	"DATA DE RETORNO",	"KW"]	
                self.df = pd.DataFrame([values], columns=colunas)

            if len(values) == 0:
                raise IndexError()

            self.MostrarDados(planilha, 'norma')    

            workbook.save('S:\\COM\Human_Resources\\01.Engineering_Tech_School\\02.Internal\\5 - Aprendizes\\3 - Manufatura\\2 - Manufatura 2023\\00 - Alunos\\Nina Cunha\\Projetos\\Instrumentos de Medição\\Instrumentos de medição.xlsx')
            workbook.close()

        except (FileNotFoundError, IndexError):
            tkinter.messagebox.showerror(title='Erro',message='Valor inválido, tentar novamente.')
    
    def buscaKW(self, op):
        try:
            self.index = []
            values = []
            kw = self.kw.get()
            kw = kw.split('/')
            ano = int(kw[1])
            semana = int(kw[0])

            workbook = openpyxl.load_workbook(filename='Instrumentos de medição.xlsx')

            worksheet = workbook['Geral']
            for col in worksheet.iter_cols(min_col = 1, max_col = 8, min_row= 1, max_row= 1):
                for col_cell in col:
                    if(col_cell.value == "KW"):
                        for row in worksheet.iter_rows(min_col = col_cell.column, max_col = col_cell.column, min_row = 2, max_row = worksheet.max_row):
                            for row_cell in row:
                                if row_cell.value is not None:
                                    if "/" in row_cell.value:
                                        kw = row_cell.value.split('/')
                                        semanaKw = int(kw[0])
                                        anoKw = int(kw[1])
                                        if semanaKw == semana and anoKw == ano:
                                            self.index.append(row_cell.row)
                                            for print_row in worksheet.iter_cols(min_col=worksheet.min_column,
                                                                                        max_col=worksheet.max_column,
                                                                                        min_row=row_cell.row,
                                                                                        max_row=row_cell.row):
                                                for row in print_row:
                                                        values.append(row.value)

            def splitList(lst, chunk_size):
                return [lst[i:i + chunk_size] for i in range(0, len(lst), chunk_size)]

            if len(values) == 0:
                raise IndexError()

            colunas = ["LOCALIZAÇÃO", "LETRA",	"DENOMINAÇÃO",	"MARCA", "NORMA/SÉRIE",	"DATA DE RETIRADA",	"DATA DE RETORNO",	"KW"]	
            self.df = pd.DataFrame(splitList(values, 8), columns=colunas)
            self.MostrarDados(['Instrumentos'], op)
        except (FileNotFoundError, IndexError):
            tkinter.messagebox.showerror(title='Erro',message='Valor inválido, tentar novamente.')

    def buscaInstrumento(self):
        values = []
        instru = self.instru.get()
        instrumento = unidecode(instru).lower()
        
        try:
            workbook = openpyxl.load_workbook(filename='Instrumentos de medição.xlsx')
            worksheet = workbook['Geral']
            for col in worksheet.iter_cols(min_col = 1, max_col = 8, min_row= 1, max_row= 1):
                for col_cell in col:
                        if(col_cell.value == "DENOMINAÇÃO"):
                            for row in worksheet.iter_rows(min_col = col_cell.column, max_col = col_cell.column,min_row = 2, max_row = worksheet.max_row):
                                for row_cell in row:
                                    if type(row_cell.value) == str:
                                        nome = row_cell.value
                                        nome = unidecode(nome).lower()
                                        if instrumento in nome:
                                            for print_row in worksheet.iter_cols(min_col=worksheet.min_column,
                                                                                max_col=worksheet.max_column,
                                                                                min_row=row_cell.row,
                                                                                max_row=row_cell.row):
                                                for row in print_row:
                                                    values.append(row.value)
                                        break
                            break
            
            def splitList(lst, chunk_size):
                return [lst[i:i + chunk_size] for i in range(0, len(lst), chunk_size)]

            colunas = ["LOCALIZAÇÃO", "LETRA",	"DENOMINAÇÃO",	"MARCA", "NORMA/SÉRIE",	"DATA DE RETIRADA",	"DATA DE RETORNO",	"KW"]	
            self.df = pd.DataFrame(splitList(values, 8), columns=colunas)

            if len(values) == 0:
                raise IndexError()

            self.MostrarDados([instru], 'instru')   

            workbook.save('S:\\COM\Human_Resources\\01.Engineering_Tech_School\\02.Internal\\5 - Aprendizes\\3 - Manufatura\\2 - Manufatura 2023\\00 - Alunos\\Nina Cunha\\Projetos\\Instrumentos de Medição\\Instrumentos de medição.xlsx')
            workbook.close()

        except (FileNotFoundError, IndexError):
            tkinter.messagebox.showerror(title='Erro',message='Valor inválido, tentar novamente.')
    
    def buscaIndex(self):
        try:
            values = []
            
            workbook = openpyxl.load_workbook(filename='Instrumentos de medição.xlsx')
            worksheet = workbook['Geral']
            
            for i in self.index:
                values.append([cell.value for cell in list(worksheet.rows)[i-1]])

            def splitList(lst, chunk_size):
                return [lst[i:i + chunk_size] for i in range(0, len(lst), chunk_size)]

            colunas = ["LOCALIZAÇÃO", "LETRA",	"DENOMINAÇÃO",	"MARCA", "NORMA/SÉRIE",	"DATA DE RETIRADA",	"DATA DE RETORNO",	"KW"]	
            self.df = pd.DataFrame(values, columns=colunas)
            self.MostrarDados(['Instrumentos'], 'kw')   

            workbook.save('S:\\COM\Human_Resources\\01.Engineering_Tech_School\\02.Internal\\5 - Aprendizes\\3 - Manufatura\\2 - Manufatura 2023\\00 - Alunos\\Nina Cunha\\Projetos\\Instrumentos de Medição\\Instrumentos de medição.xlsx')
            workbook.close()

            if values is None:
                raise IndexError()
        except (FileNotFoundError, IndexError):
            tkinter.messagebox.showerror(title='Erro',message='Valor inválido, tentar novamente.')

    

    def MostrarDados(self, planilha, op):
        self.limpar()
        
        title = tkinter.Label(self.master, text=planilha[0], font=("bold", 16))
        title.pack(fill=tkinter.X, pady=10)
        self.selecao = []
        selecaoId = []

        tree = tkinter.ttk.Treeview(self.master, columns=list(self.df.columns), show='headings')
        tree.pack(fill=tkinter.BOTH, expand=True)
        tree.master.minsize(width=900, height=200)
        tree.master.maxsize(width=1300, height=500)
        tree.master.resizable(width=None, height=None)

        def select(_):
            item = self.tree.item(self.tree.selection())
            if item['values'] not in self.selecao:
                self.selecao.append(item['values'])

        def delete(_):
            item = self.tree.item(self.tree.selection())
            if item['values'] in self.selecao:
                self.selecao.remove(item['values'])


        tree.bind('<<TreeviewSelected>>', select)
        tree.bind('<Double-1>', delete)
        
        for col in self.df.columns:
            tree.heading(col, text=col)
            tree.column(col, width=150)

        for _, row in self.df.iterrows():
            tree.insert("", "end", values=list(row))

        def botKW():
            novoKW = tkinter.simpledialog.askstring("Input", "Novo KW:")
            try:
                kw = novoKW
                kw = kw.split('/')

                ano = int(kw[1])
                semana = int(kw[0])
                
                if semana > 52 or semana < 1  or ano >= 100:
                    raise ValueError()
                else:
                    print(self.selecao)
                    if len(self.selecao) == 0:
                        self.mudarKW(self.df.iloc[0].values[-1], novoKW)
                    else:
                        print('entrou')
                        self.mudarKWSelected(self.df.iloc[0].values[-1], novoKW)
                    self.mudarRetorno()
            except (ValueError, IndexError):
                tkinter.messagebox.showerror(title='Erro',message='Valor inválido, tentar novamente.')
            
        def botRetirada():
            self.mudarRetirada()

        botao_modkw = tkinter.Button(text='Modificar KW', command=botKW)
        botao_modkw.pack(side="right", padx=250)

        botao_mod = tkinter.Button(text='Modificar Retirada', command=botRetirada)
        botao_mod.pack(side="left", padx=250)

        botao_voltar = tkinter.Button(text='Voltar',command=self.paginaInicial)
        botao_voltar.place(relx=0.005, rely=0.005, anchor='nw')


    def mudarKWSelected(self, kw, novoKW):
        try:
            workbook = openpyxl.load_workbook(filename='Instrumentos de medição.xlsx')
            worksheet = workbook['Geral']

            values = [self.selecao[i][0] for i in range(len(self.selecao))]
            print(f"Linhas selecionadas {values}")
            for row in worksheet.iter_rows:
                if row[0] in values:
                    for col in worksheet.iter_cols(min_col = 1, max_col = 8, min_row= 1, max_row= 1):
                        for col_cell in col:
                            if(col_cell.value == "KW"):
                                for i in self.index:
                                    for row in worksheet[i]:
                                            if kw in str(row.value):
                                                row.value = novoKW
                                                workbook.save('Instrumentos de medição.xlsx')
        except:
            tkinter.messagebox.showerror(title='Erro',message='Algo deu errado, tentar novamente.')


    def mudarKW(self, kw, novoKW):
        try:

            workbook = openpyxl.load_workbook(filename='Instrumentos de medição.xlsx')
            worksheet = workbook['Geral']
            for col in worksheet.iter_cols(min_col = 1, max_col = 8, min_row= 1, max_row= 1):
                for col_cell in col:
                    if(col_cell.value == "KW"):
                        for i in self.index:
                            for row in worksheet[i]:
                                    if kw in str(row.value):
                                        row.value = novoKW
                                        workbook.save('Instrumentos de medição.xlsx')
        except:
            tkinter.messagebox.showerror(title='Erro',message='Algo deu errado, tentar novamente.')
    
    def mudarRetorno(self):
        try:
            workbook = openpyxl.load_workbook(filename='Instrumentos de medição.xlsx')
            worksheet = workbook['Geral']
            for col in worksheet.iter_cols(min_col = 1, max_col = 8, min_row= 1, max_row= 1):
                for col_cell in col:
                    if(col_cell.value == "DATA DE RETORNO"):
                        for i in self.index:
                            c = 0
                            for row in worksheet[i]:
                                c+=1
                                if c == 7:
                                    hoje = datetime.datetime.now()
                                    row.value = hoje.strftime("%d-%m-%Y %H:%M:%S")
                                    workbook.save('Instrumentos de medição.xlsx')
            tkinter.messagebox.showinfo(title='Operação Finalizada',message='A operação foi executada corretamente.')
            self.buscaIndex()
        except:
            tkinter.messagebox.showerror(title='Erro',message='Algo deu errado, tentar novamente.')

    def mudarRetirada(self):
        try:
            workbook = openpyxl.load_workbook(filename='Instrumentos de medição.xlsx')
            worksheet = workbook['Geral']
            max_row = worksheet.max_row
            for col in worksheet.iter_cols(min_col = 1, max_col = 8, min_row= 1, max_row= 1):
                for col_cell in col:
                    if(col_cell.value == "DATA DE RETIRADA"):
                        for i in self.index:
                            c = 0
                            for row in worksheet[i]:
                                c+=1
                                if c == 6:
                                    hoje = datetime.datetime.now()
                                    row.value = hoje.strftime("%d-%m-%Y %H:%M:%S")
                                    workbook.save('Instrumentos de medição.xlsx')
            tkinter.messagebox.showinfo(title='Operação Finalizada',message='A operação foi executada corretamente.')
            self.buscaIndex()
        except:
            tkinter.messagebox.showerror(title='Erro',message='Algo deu errado, tentar novamente.')
    

root = tkinter.Tk()
app = App(root)
app.mainloop()

# 4738440068- 8070208 (teste de perdidos)
# 4738310040-724040 (teste sucateados)
# 4737910007- 750562 (teste geral)