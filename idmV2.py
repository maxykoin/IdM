import tkinter as tk
from tkinter import ttk, simpledialog, filedialog   
import ttkbootstrap as ttk 
import pandas as pd
from openpyxl import load_workbook
from datetime import datetime
from unidecode import unidecode

class App(tk.Frame):
    def __init__(self, master):
        super().__init__(master)
        self.master.title("Instrumentos de Medição - ETS")
        self.master.geometry('800x800')
        self.pagInicial()

    def limpar(self):
        for widget in self.master.winfo_children():
            widget.destroy()
    
    def buscaPlanilha(self):
        try:
            self.df = pd.read_excel('Instrumentos de medição.xlsx')
            self.planilha = load_workbook("Instrumentos de medição.xlsx")
            self.geral = self.planilha['Geral']
            self.perdido = self.planilha['Perdidos']
        except:
            tk.messagebox.showerror(title='Erro', message="Base de dados não encontrada.")
        else:
            self.buscaDados()
    
    def pagInicial(self):
        self.limpar()
        
        frame = tk.Frame(root, width=200, height=200)
        frame.pack(expand=True)   

        label = tk.Label(frame, text="Instrumentos de Medição", font = 'Calibri 24 bold')
        label.pack()
        
        tef = tk.Label(frame, text="Pesquise por:", font = 'Calibri 18')
        tef.pack(pady=20)

        def inputBox(op):
            cracha = tk.Toplevel(root)

            width = self.master.winfo_screenwidth()
            height = self.master.winfo_screenheight()
            left = int(width / 2 - 200 / 2)
            top = int(height / 2 - 200 /2)

            cracha.geometry(f'200x200+{left}+{top}')

            label = tk.Label(cracha, font = 'Calibri 18 bold')
            label.pack(pady=5)

            self.entry = tk.Entry(cracha)
            self.entry.pack(pady=5)
            self.entry.focus_set()
            self.entry.bind('<Return>', lambda event: treat())

            if op == 'norma':
                label['text'] = 'Insira a Norma:'
                self.op = 'norma'
            elif op == 'kw':
                label['text'] = 'Insira a KW:'
                self.op = 'kw'
            else:
                label['text'] = 'Insira o Nome'
                self.op = 'nome'

            def treat():
                try: 
                    if self.op == 'kw':
                        kw = self.entry.get().split('/')
                        semana = int(kw[0])
                        ano = int(kw[1])
                        if '/' not in self.entry.get():
                            raise ValueError
                        elif semana > 52 or semana < 1 or ano > 99 or ano < 10:
                            raise ValueError
                    elif self.op == 'norma':
                        self.norma = self.entry.get()
                        self.final = self.entry.get().split('-')
                        self.final = self.final[-1].strip()
                        self.norma = ''.join(self.norma.split())
                    elif self.op == 'nome':
                        self.nome = unidecode(self.entry.get()).lower()
                    if not self.entry.get():
                        raise ValueError
                except ValueError:
                    tk.messagebox.showerror(title='Erro', message="Insira um valor valido.")
                    inputBox(self.op)
                else:
                    self.buscaPlanilha()

            butBusca = tk.Button(cracha, text="Pesquisar", font='Calibri 12', width=10 , height=1, command=treat)
            butBusca.pack(pady=15)

        butFrame = tk.Frame(frame)
        butFrame.pack()

        butNorma = tk.Button(butFrame, text="Norma", font='Calibri 12', width=10 , height=2, command=lambda:inputBox('norma'))
        butNorma.pack(side=tk.LEFT, padx=10)

        butKw = tk.Button(butFrame, text="KW", font='Calibri 12', width=10 , height=2,command=lambda:inputBox('kw'))
        butKw.pack(side=tk.LEFT, padx=10)

        butNome = tk.Button(butFrame, text="Nome", font='Calibri 12', width=10 , height=2,command=lambda:inputBox('nome'))
        butNome.pack(side=tk.LEFT, padx=10)
    
    def buscaDados(self):
        dados = []
        try:
            for sheet in self.planilha.worksheets:
                sheet = self.planilha[sheet.title]
                columns = [cell.value for cell in next(sheet.iter_rows())]
                for row in sheet.iter_rows():
                    for i in range(len(row)):
                        if row[i].value in [self.entry.get()]:
                            item = [cell.value for cell in row]
                            if not any(item[:3] == data[:3] for data in dados):
                                dados.append(item)
                                self.sheet = sheet.title
                        if self.op == 'nome' and i == 2:
                            instrumento = unidecode(str(row[i].value)).lower()
                            if self.nome in instrumento.split():
                                item = [cell.value for cell in row]
                                if not any(item[:3] == data[:3] for data in dados):
                                    dados.append(item)
                                    self.sheet = sheet.title
                        if self.op == 'norma' and i == 4:
                            instrumento = str(row[i].value)
                            instrumento = instrumento.strip()
                            final = instrumento.split('-')
                            final = final[-1].strip()
                            instrumento = ''.join(instrumento.split())
                            if final in [self.final] or instrumento in [self.norma]:
                                item = [cell.value for cell in row]
                                if not any(item[:3] == data[:3] for data in dados):
                                    dados.append(item)
                                    self.sheet = sheet.title

            self.df = pd.DataFrame(dados, columns=columns)
        except:
            tk.messagebox.showerror(title='Erro', message="Valor não encontrado.")
        else:
            self.mostraDados()

    def mostraDados(self):
        self.limpar()
        self.selecionado = []

        voltarButton = tk.Button(root, text="  ←  ", command=self.pagInicial)
        voltarButton.pack(anchor='nw', padx=10, pady=10)

        label = tk.Label(root, text=self.sheet, font = 'Calibri 18 bold')
        label.pack()

        paragraph = tk.Label(root, text="Clique 1 vez para selecionar \n Clique 2 vezes para remover", font = 'Calibri 12')
        paragraph.pack(pady=20)
        
        self.colunas = self.df.columns.tolist()
        dados = self.df.values.tolist() 

        self.table = ttk.Treeview(root, columns=self.colunas, show = 'headings')
        [self.table.heading(i, text=i) for i in self.colunas]
        [self.table.insert(parent='', index= 0, values=i)  for i in dados]
        self.table.tag_configure('selected', background='light grey')
        self.table.pack(fill='both', expand=True)

        def select(_):
            self.table.item(self.table.selection(), tags=('selected',))
            item = self.table.item(self.table.selection())
            if item['values'] not in self.selecionado:
                self.selecionado.append(item['values'])

        def delete(_):
            self.table.item(self.table.selection(), tags=())
            item = self.table.item(self.table.selection())
            if item['values'] in self.selecionado:
                self.selecionado.remove(item['values'])
        
        def novoKw():
            cracha = tk.Toplevel(root)

            width = self.master.winfo_screenwidth()
            height = self.master.winfo_screenheight()
            left = int(width / 2 - 200 / 2)
            top = int(height / 2 - 200 /2)

            cracha.geometry(f'200x200+{left}+{top}')

            label = tk.Label(cracha, text='Insira um novo kw', font = 'Calibri 18 bold')
            label.pack(pady=5)

            self.novokw = tk.Entry(cracha)
            self.novokw.pack(pady=5)
            self.novokw.focus_set()
            self.novokw.bind('<Return>', lambda event: self.mudaKw())

            butBusca = tk.Button(cracha, text="Pesquisar", font='Calibri 12', width=10 , height=1, command=self.mudaKw)
            butBusca.pack(pady=15)

        def novaRetirada():
            try:
                if len(self.selecionado) == 0:
                    for row in self.planilha[self.sheet].iter_rows():
                        if any(row[4].value == item[4] for item in self.df.values):
                            hoje = datetime.now()
                            row[5].value = hoje.strftime("%d-%m-%Y %H:%M:%S")
                else:
                    for row in self.planilha[self.sheet].iter_rows():
                        if any(row[4].value == selected[4] for selected in self.selecionado):
                            hoje = datetime.now()
                            row[5].value = hoje.strftime("%d-%m-%Y %H:%M:%S")

                self.planilha.save('Instrumentos de medição.xlsx')
                tk.messagebox.showinfo(title='Sucesso', message="Valor modificado com sucesso.")
            except:
                tk.messagebox.showerror(title='Erro', message='Erro')

        self.table.bind('<<TreeviewSelect>>', select)
        self.table.bind('<Double-1>', delete)

        butFrame = tk.Frame()
        butFrame.pack(side=tk.BOTTOM)

        butKw = tk.Button(butFrame, text="Mudar KW", font = 'Calibri 12', width=15, height=1, command=lambda: novoKw())
        butKw.pack(side=tk.LEFT, pady=10)
        
        butRetirada = tk.Button(butFrame, text="Mudar Retirada", font = 'Calibri 12', width=15, height=1, command=lambda: novaRetirada())
        butRetirada.pack(padx=25, pady=10)

    def mudaKw(self):
        try:
            kw = self.novokw.get().split('/')
            semana = int(kw[0])
            ano = int(kw[1])

            if len(kw) != 2 or semana < 1 or semana > 52 or ano < 10 or ano > 99:
                raise ValueError("Invalid KW format.")

            if len(self.selecionado) == 0:
                for row in self.planilha[self.sheet].iter_rows():
                    if any(row[4].value == item[4] for item in self.df.values):
                        row[7].value = self.novokw.get()
                        hoje = datetime.now()
                        row[6].value = hoje.strftime("%d-%m-%Y %H:%M:%S")
            else:
                for row in self.planilha[self.sheet].iter_rows():
                    if any(row[4].value == selected[4] for selected in self.selecionado):
                        row[7].value = self.novokw.get()
                        hoje = datetime.now()
                        row[6].value = hoje.strftime("%d-%m-%Y %H:%M:%S")

            self.planilha.save('Instrumentos de medição.xlsx')
            tk.messagebox.showinfo(title='Sucesso', message="Valor modificado com sucesso.")
        
        except ValueError as ve:
            tk.messagebox.showerror(title='Erro', message=str(ve))

root = ttk.Window(themename ='pulse')
app = App(root)
root.mainloop()